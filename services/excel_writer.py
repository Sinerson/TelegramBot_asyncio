import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook


async def export_to_excel(data: dict, file_name: str):
    wb = Workbook()
    ws = wb.active
    df = pd.DataFrame([data])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
    wb.save(f"{file_name}.xlsx")
